import os
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

METRICS = ["maxvout", "rise_delay", "fall_delay", "rise_trans", "fall_trans", "iddq_leak"]

def parse_mt0_file(filepath):
    maxvout_ref = None
    header_idx  = None
    headers     = None
    with open(filepath, 'r', errors='ignore') as f:
        for i, line in enumerate(f):
            if line.strip().lower().startswith('.title'):
                m = re.search(r'(\d+[\.pP]\d+)[vV]', line)
                if m:
                    maxvout_ref = float(m.group(1).lower().replace('p', '.'))

            if line.strip().lower().startswith('index'):
                headers    = line.strip().split()
                header_idx = i
                break                       

    if header_idx is None or headers is None:
        return maxvout_ref, {}

    metrics_in_file = [m for m in METRICS if m in headers]
    if not metrics_in_file:
        return maxvout_ref, {}


    try:
        df = pd.read_csv(
            filepath,
            sep        = r'\s+',
            skiprows   = header_idx,
            names      = headers,
            header     = 0,
            usecols    = metrics_in_file,
            dtype      = float,             
            na_values  = ['.', 'NA', ''],
            on_bad_lines = 'skip',
            engine     = 'c',
        )
    except Exception:
        return maxvout_ref, {}

    df.dropna(inplace=True)
    data = {m: df[m].to_numpy() for m in metrics_in_file}  
    for m in METRICS:
        data.setdefault(m, np.array([]))

    return maxvout_ref, data


def calculate_yield(metric, arr, maxvout_reference=None):
    if arr is None or len(arr) == 0:
        return None

    if metric == "maxvout":
        if maxvout_reference is None:
            return None
        threshold = maxvout_reference * 0.80       
        failed    = int(np.sum(arr < threshold))    
        return (failed / len(arr)) * 100

    else:
        threshold = arr[0] * 1.20                  
        remaining = arr[1:]                        
        if len(remaining) == 0:
            return None
        failed = int(np.sum(remaining > threshold)) # vectorized
        return (failed / len(remaining)) * 100


def main():
    BASE_PATH = input("Enter input folder path : ").strip().strip('"')

    if not os.path.isdir(BASE_PATH):
        print(f"ERROR: Input path does not exist: {BASE_PATH}")
        return

    file_name = input("Enter output file name  : ").strip().strip('"')

    if not file_name.lower().endswith('.xlsx'):
        file_name += '.xlsx'

    OUTPUT_PATH = os.path.join(BASE_PATH, file_name)

    print(f"\nScanning : {BASE_PATH}")
    print(f"Output   : {OUTPUT_PATH}\n")

    results = {m: [] for m in METRICS}

    for cell_name in sorted(os.listdir(BASE_PATH)):
        cell_path = os.path.join(BASE_PATH, cell_name)
        if not os.path.isdir(cell_path):
            continue

        for category in sorted(os.listdir(cell_path)):
            cat_path = os.path.join(cell_path, category)
            if not os.path.isdir(cat_path):
                continue

            mt0_files = [f for f in os.listdir(cat_path) if f.lower().endswith('.mt0')]

            if not mt0_files:
                print(f"  [{cell_name} / {category}] no .mt0 — yield left valueless")
                for metric in METRICS:
                    results[metric].append({
                        "Cell_name"  : cell_name,
                        "Category"   : category,
                        "Yield(%)": "mt0 file doesnot exits"
                    })
                continue

            mt0_path = os.path.join(cat_path, mt0_files[0])
            print(f"  [{cell_name} / {category}] {mt0_files[0]}")
            maxvout_ref, data = parse_mt0_file(mt0_path)

            for metric in METRICS:
                arr = data.get(metric, np.array([]))
                if len(arr) == 0:
                    yld = ""
                else:
                    yld = calculate_yield(metric, arr, maxvout_ref)
                    yld = round(yld, 4) if yld is not None else ""

                results[metric].append({
                    "Cell_name"  : cell_name,
                    "Category"   : category,
                    "Yield(%)": yld
                })

  
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font  = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill  = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    header_align = Alignment(horizontal="center")
    data_font    = Font(name="Arial")
    empty_font   = Font(name="Arial", color="AAAAAA", italic=True)

    for metric in METRICS:
        ws = wb.create_sheet(title=metric)
        ws.append(["Cell_name", "Category", "Yield(%)"])

        for col_idx in range(1, 4):
            cell           = ws.cell(row=1, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

        prev_cell_name = None
        for row in results[metric]:
            cell_display   = row["Cell_name"] if row["Cell_name"] != prev_cell_name else ""
            prev_cell_name = row["Cell_name"]

            ws.append([cell_display, row["Category"], row["Yield(%)"]])
            r = ws.max_row
            for col_idx in range(1, 4):
                ws.cell(row=r, column=col_idx).font = data_font

            if row["Yield(%)"] == "":
                yc       = ws.cell(row=r, column=3)
                yc.value = "-"
                yc.font  = empty_font

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 15

    wb.save(OUTPUT_PATH)
    print(f"\nSaved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()