import pandas as pd#ff4e4e
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


def read_excel_file(file):
    df = pd.read_excel(file)
    cell_data1 = {}
    cell_data2 = {}
    condition = input("Enter condition (NOM/+-NOM/+-Value): ")

    for index, row in df.iterrows():
        if pd.notna(row["Parameter"]):
            parts = row["Parameter"].split('@')
            if parts[-1].upper() == condition.upper():
                left       = parts[0]
                name       = parts[0].split("_", 1)[1].split("~")[0]
                param_type = left.split("~")[1]
                if name not in cell_data1:
                    cell_data1[name] = {'SLICE_NAME': None, 'FREQ': 0, 'IDDQ': 0, 'IDDA': 0}
                    if len(parts) > 1:
                        cell_data1[name]['SLICE_NAME'] = parts[0].split('_', 1)[0]

                if name not in cell_data2 and cell_data1[name].get(param_type, 0) != 0:
                    cell_data2[name] = {'SLICE_NAME': None, 'FREQ': 0, 'IDDQ': 0, 'IDDA': 0}
                    if len(parts) > 1:
                        cell_data2[name]['SLICE_NAME'] = parts[0].split('_', 1)[0]

               
                if param_type in ("FREQ", "IDDQ", "IDDA"):
                    value = row["Median"] / 1000 if param_type == "FREQ" else row["Median"]
                    if cell_data1[name][param_type] == 0:
                        cell_data1[name][param_type] = value
                    elif name in cell_data2 and cell_data2[name][param_type] == 0:
                        cell_data2[name][param_type] = value

    df1 = pd.DataFrame.from_dict(cell_data1, orient="index")
    df2 = pd.DataFrame.from_dict(cell_data2, orient="index")
    df1.index.name = "Cell Name"
    df2.index.name = "Cell Name"
    df1.to_excel("cell_silicon_report16.xlsx", index_label="Cell Name")
    df2.to_excel("cell_silicon_report17.xlsx", index_label="Cell Name")
    df1 = df1.reset_index()
    df2 = df2.reset_index()
    return df1, df2


def comparison_silicon_vs_spice(file1, file2):
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2,
        usecols=["Slice Name", "VDD Ring", "Cell Name",
                 "FREQ[KHz][FREQ/(1000*1024)]", "IDDQ [A]", "[IDDA - IDDQ] Diff"]
    )
    df2 = df2.rename(columns={
        "FREQ[KHz][FREQ/(1000*1024)]": "FREQ",
        "[IDDA - IDDQ] Diff":          "IDDA",
        "IDDQ [A]":                    "IDDQ",
    })

    df = pd.merge(df1, df2, on="Cell Name", suffixes=("_silicon", "_spice"))

    df["FREQ_Diff%"] = (abs(df["FREQ_silicon"] - df["FREQ_spice"]) / df["FREQ_spice"]) * 100
    df["IDDA_Diff%"] = (abs(df["IDDA_silicon"] - df["IDDA_spice"]) / df["IDDA_spice"]) * 100
    df["IDDQ_Diff%"] = (abs(df["IDDQ_silicon"] - df["IDDQ_spice"]) / df["IDDQ_spice"]) * 100

    THRESHOLDS = {"FREQ": 5.0, "IDDQ": 30.0, "IDDA": 30.0}

    columns = pd.MultiIndex.from_tuples([
        ("Slice Name", ""),
        ("VDD Ring", ""),
        ("Cell Name", ""),
        ("SPICE",   "FREQ"),
        ("SPICE",   "IDDA"),
        ("SPICE",   "IDDQ"),
        ("SILICON", "FREQ"),
        ("SILICON", "IDDA"),
        ("SILICON", "IDDQ"),
        ("DIFF%",   "FREQ"),
        ("DIFF%",   "IDDA"),
        ("DIFF%",   "IDDQ"),
    ])

    data = pd.DataFrame(columns=columns)
    data[("Slice Name","")] = df["Slice Name"].values
    data[("VDD Ring", "")]   = df["VDD Ring"].values
    data[("Cell Name", "")]  = df["Cell Name"].values
    data[("SPICE",   "FREQ")]       = df["FREQ_spice"].values
    data[("SPICE",   "IDDA")]       = df["IDDA_spice"].values
    data[("SPICE",   "IDDQ")]       = df["IDDQ_spice"].values
    data[("SILICON", "FREQ")]       = df["FREQ_silicon"].values
    data[("SILICON", "IDDA")]       = df["IDDA_silicon"].values
    data[("SILICON", "IDDQ")]       = df["IDDQ_silicon"].values
    data[("DIFF%",   "FREQ")]       = df["FREQ_Diff%"].values
    data[("DIFF%",   "IDDA")]       = df["IDDA_Diff%"].values
    data[("DIFF%",   "IDDQ")]       = df["IDDQ_Diff%"].values

    output = "comparison_report2.xlsx"
    wb = load_workbook.__module__ 
    from openpyxl import Workbook 
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    top_levels   = [col[0] for col in data.columns]  # top level header like SPICE/SILICON
    sub_levels   = [col[1] for col in data.columns]  # sub-level headers like FREQ/IDDQ/IDDA
    col_offset = 1
    i = 0
    while i < len(top_levels):
        group = top_levels[i]
        span = 1
        while i + span < len(top_levels) and top_levels[i + span] == group:
            span += 1
        start_col = col_offset + i
        end_col   = start_col + span - 1
        ws.cell(row=1, column=start_col, value=group)
        if span > 1:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1,   end_column=end_col
            )
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center")
        i += span

    
    for j, sub in enumerate(sub_levels, start=1):
        ws.cell(row=2, column=j, value=sub)

   
    for r_idx, row_vals in enumerate(data.itertuples(index=False), start=3):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

   
    diff_cols = {
        col_idx: sub
        for col_idx, (top, sub) in enumerate(zip(top_levels, sub_levels), start=1)
        if top == "DIFF%"
    }

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.column in diff_cols:
                sub_col   = diff_cols[cell.column]
                threshold = THRESHOLDS.get(sub_col, 5.0)
                cell.fill = RED_FILL if abs(float(cell.value)) > threshold else GREEN_FILL
                

    wb.save(output)
    return output

if __name__ == '__main__':

    read_excel_file('8XYA12376.000_wafers.xlsx')
    print("Enter a report number you want 1 or 2: ")
    press = int(input())
    if press == 1:
        comparison_silicon_vs_spice(
            "cell_silicon_report16.xlsx",
            "V1.0_3.1_sabdbox_uhvt_rev2_2_20240223_V0p8_TT25.xlsx")
    if press == 2:
        comparison_silicon_vs_spice(
            "cell_silicon_report17.xlsx",
            "V1.0_3.1_sabdbox_uhvt_rev2_2_20240223_V0p8_TT25.xlsx")